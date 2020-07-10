from fbs_runtime.application_context import ApplicationContext

# -*- coding: utf-8 -*-

import sys, xlrd, csv
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QDialog, QApplication
from io import StringIO
import pandas as pd
from PandasModel import PandasModel
from MeatPart_Tool_V4 import Application # import du code MeatPart_Tool_V4.py (interface graphique)
import csv # import du module csv permettant de gérer les données csv
from Division_Database import preparation_bdd # import du code Division_Database (mise en forme de la bdd pour pouvoir appliquer la fonction calcul)
from Fonction_Calcul_Bovin import calcul_bovin # import du code Fonction_Calcul (permet de calculer les facteurs d'allocation biophysique, massique et économique)
from Fonction_Calcul_Ovin import calcul_ovin # import du code Fonction_Calcul (permet de calculer les facteurs d'allocation biophysique, massique et économique)
from Fonction_Calcul_Veau import calcul_veau # import du code Fonction_Calcul (permet de calculer les facteurs d'allocation biophysique, massique et économique)
from test_float import isfloat # import du code test_float permettant de tester si une variable est convertible en float
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.chart import BarChart, Reference, Series
from PyQt5.QtWidgets import QApplication
from errors import Title, Text, Details
import os



app2 = QApplication([])
screen_resolution2 = app2.desktop().screenGeometry()
width2 = screen_resolution2.width()
height2 = screen_resolution2.height()
ratio_largeur2 = width2/1920
ratio_hauteur2 = height2/1200

class AppWindow(QDialog):

    def __init__(self):
        super().__init__()
        self.ui = Application()
        self.ui.setupUi(self)
        self.show()
        self.ui.pushButton.clicked.connect(self.Chargement) # un clic sur le bouton appliquera la méthode Chargement
        self.ui.combobox_especes_tab_2.currentTextChanged.connect(self.Combobox_especes_tab_2)
        self.ui.combobox2.currentTextChanged.connect(self.Combobox2)
        self.ui.combobox3.currentTextChanged.connect(self.Combobox3)
        self.ui.combobox4.currentTextChanged.connect(self.Combobox4)
        self.ui.combobox5.currentTextChanged.connect(self.Combobox5)
        self.ui.combobox6.currentTextChanged.connect(self.Combobox6)
        self.ui.combobox7.currentTextChanged.connect(self.Combobox7)
        self.ui.Bouton_Valider.clicked.connect(self.Ajouter) # un clic sur le bouton appliquera la méthode Ajouter
        self.ui.Bouton_Modele.clicked.connect(self.Modele) # un clic sur le bouton Modele appliquera la méthode Modele
        self.ui.Bouton_Effacer.clicked.connect(self.Effacer) # un clic sur le bouton Effacer appliquera la méthode Effacer
        self.ui.Bouton_Excel.clicked.connect(self.Excel) # un clic sur le bouton Excel appliquera la méthode Excel
        #self.ui.bouton_tuto.clicked.connect(self.Tuto)
        #self.ui.bouton_rapport.clicked.connect(self.Rapport)


    #def Tuto(self):
        #os.system('D:/Professionnel/Recherche/2019/Interbev/Logos/TUTO.pdf')

    #def Rapport(self):
        #os.system('D:/Professionnel/Recherche/2019/Interbev/Logos/RAPPORT.pdf')

    def Chargement(self):
        fileName, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Open File", "", "CSV Files (*.csv)");
        self.ui.pathLE.setText(fileName)

        # Ouverture de la bdd avec Pandas pour l'affichage
        df = pd.read_csv(fileName, encoding = "utf-8",delimiter=';')
        model = PandasModel(df)
        self.ui.pandasTv.setModel(model)

        # Ouverture de la bdd avec "csv"
        df = open(fileName, encoding = 'utf-8')
        bdd = csv.reader(df, delimiter=';')

        #preparation de la bdd
        self.liste_finale = preparation_bdd(bdd)
        liste_finale = self.liste_finale

        # On remplit les combobox des tab2 et tab4 avec la listes des espèces
        def especes():
            liste_especes = []
            for j in range(0, len(liste_finale)):
                for i in range(0, len(liste_finale[j])):
                    liste_especes.append(liste_finale[j][i][0])
            liste_especes = list(set(liste_especes))
            return liste_especes

        self.liste_especes = especes()
        self.ui.combobox_especes_tab_2.clear()
        self.ui.combobox4.clear()
        for i in range(0,len(self.liste_especes)):
            self.ui.combobox_especes_tab_2.addItem(self.liste_especes[i])
            self.ui.combobox4.addItem(self.liste_especes[i])

    # On remplit la combobox2 en fonction de la valeur sélectionnée dans la combobox1
    def Combobox_especes_tab_2(self):
        self.ui.combobox2.clear()
        liste_especes = self.liste_especes
        liste_finale = self.liste_finale
        liste_races = []

        for i in range(0,len(liste_especes)):
            if self.ui.combobox_especes_tab_2.currentText() == liste_especes[i]:
                for k in range(0, len(liste_finale)):
                    for j in range(0, len(liste_finale[k])):
                        if liste_finale[k][j][0] == liste_especes[i]:
                            liste_races.append(liste_finale[k][j][1])
        self.liste_races = list(set(liste_races))

        for i in range(0,len(self.liste_races)):
            self.ui.combobox2.addItem(self.liste_races[i])

    # On remplit la combobox3 en fonction de la valeur sélectionnée dans la combobox2
    def Combobox2(self):
        self.ui.combobox3.clear()
        liste_especes = self.liste_especes
        liste_finale = self.liste_finale
        liste_races = self.liste_races
        liste_mois = []

        for o in range(0,len(liste_especes)):
            if self.ui.combobox_especes_tab_2.currentText() == liste_especes[o]:
                for i in range(0,len(liste_races)):
                    if self.ui.combobox2.currentText() == liste_races[i]:
                        for k in range(0, len(liste_finale)):
                            for j in range(0, len(liste_finale[k])):
                                if liste_finale[k][j][0] == liste_especes[o] and liste_finale[k][j][1] == liste_races[i]:
                                    liste_mois.append(liste_finale[k][j][2])
        self.liste_mois = list(set(liste_mois))

        for i in range(0,len(self.liste_mois)):
            self.ui.combobox3.addItem(self.liste_mois[i])


    # On remplit la combobox7 en fonction de la valeur sélectionnée dans la combobox3
    def Combobox3(self):
        self.ui.combobox7.clear()
        liste_especes = self.liste_especes
        liste_finale = self.liste_finale
        liste_races = self.liste_races
        liste_mois = self.liste_mois
        liste_elevage = []

        for o in range(0,len(liste_especes)):
            if self.ui.combobox_especes_tab_2.currentText() == liste_especes[o]:
                for p in range(0,len(liste_races)):
                    if self.ui.combobox2.currentText() == liste_races[p]:
                        for i in range(0, len(liste_mois)):
                            if self.ui.combobox3.currentText() == liste_mois[i]:
                                for k in range(0, len(liste_finale)):
                                    for j in range(0, len(liste_finale[k])):
                                        if liste_finale[k][j][0] == liste_especes[o] and liste_finale[k][j][1] == liste_races[p] and liste_finale[k][j][2] == liste_mois[i]:
                                            liste_elevage.append(liste_finale[k][j][3])
        liste_elevage = list(set(liste_elevage))

        for i in range(0,len(liste_elevage)):
            self.ui.combobox7.addItem(liste_elevage[i])


    # On remplit le tableau de l'onglet 2 en fonction de l'individu choisi (donc le choix de la combobox3)
    # En plus, on lance le calcul des facteurs d'allocation pour l'individu
    def Combobox7(self):
        # On crée une liste représentant l'individu dès qu'on change la valeur de la combobox3
        individu = []
        individu.append(self.ui.combobox_especes_tab_2.currentText())
        individu.append(self.ui.combobox2.currentText())
        individu.append(self.ui.combobox3.currentText())
        individu.append(self.ui.combobox7.currentText())

        # on affiche les infos dans l'onglet résultats
        self.ui.texte_specie_result.setText(self.ui.combobox_especes_tab_2.currentText())
        self.ui.texte_race_result.setText(self.ui.combobox2.currentText())
        self.ui.texte_category_result.setText(self.ui.combobox3.currentText())
        self.ui.texte_rearing_result.setText(self.ui.combobox7.currentText())

        # On crée une liste qui représente l'individu et ses données associées et on l'affiche dans le tableau
        data_individu = []
        liste_finale = self.liste_finale
        if self.ui.combobox7.count() != 0 :
            for i in range(0,len(liste_finale)):
                for j in range(0, len(liste_finale[i])):
                    if (liste_finale[i][j][0] == individu[0] and liste_finale[i][j][1] == individu[1]
                        and liste_finale[i][j][2] == individu[2] and  liste_finale[i][j][3] == individu[3]):
                        data_individu.append(liste_finale[i][j])
                        for k in range(0,len(data_individu)):
                            for l in range(0,len(data_individu[0])):
                                valeur_numerique_defaut = QtWidgets.QTableWidgetItem(str(data_individu[k][l]))
                                self.ui.tableIndividu.setItem(k,l,valeur_numerique_defaut)

            # On lance le programme de calcul pour la liste
            if data_individu[0][0] == 'Bovine':
                Alloc_fact = calcul_bovin(data_individu)
            if data_individu[0][0] == 'Ovine':
                Alloc_fact = calcul_ovin(data_individu)
            if data_individu[0][0] == 'Calf':
                Alloc_fact = calcul_veau(data_individu)

            self.Biophys_alloc_fact = Alloc_fact[0]
            self.Mass_alloc_fact = Alloc_fact[1]
            self.Eco_alloc_fact = Alloc_fact[2]
            self.Biophys_alloc_fact_kg = Alloc_fact[3]
            self.Mass_alloc_fact_kg = Alloc_fact[4]
            self.Eco_alloc_fact_kg = Alloc_fact[5]

            self.data_individu = data_individu

        # On affiche les facteurs d'allocation dans le tableau de l'onglet 3
        facteurs = []
        for i in range (0,len(data_individu)):
            facteurs.append([])
            facteurs[i].append(data_individu[i][4])
            facteurs[i].append(data_individu[i][5])
            facteurs[i].append("%.4f" % self.Biophys_alloc_fact[i])
            facteurs[i].append("%.4f" % self.Mass_alloc_fact[i])
            facteurs[i].append("%.4f" % self.Eco_alloc_fact[i])
            facteurs[i].append("%.4f" % self.Biophys_alloc_fact_kg[i])
            facteurs[i].append("%.4f" % self.Mass_alloc_fact_kg[i])
            facteurs[i].append("%.4f" % self.Eco_alloc_fact_kg[i])

        for k in range(0,len(facteurs)):
            for l in range(0,len(facteurs[0])):
                valeur_numerique_defaut = QtWidgets.QTableWidgetItem(str(facteurs[k][l]))
                self.ui.table_facteurs.setItem(k,l,valeur_numerique_defaut)

    # Préparation d'une sous-liste (classement par destination) pour tracer des graphes
        liste_destination = ['Pet Food', 'PAP C3', 'Gelatin C3', 'C1-C2 for disposal', 'Skin tannery C3', 'Human food', 'Fat and greaves C3', 'Spreading/Compost']

        # Pour chaque destination, on construit une liste qui comprend les 3 facteurs d'allocation (dans l'ordre biophysique, massique, économique)
        alloc_fact_tri = []
        Pet_food_facteur = [0,0,0]
        PAP_facteur = [0,0,0]
        Gelatin_facteur = [0,0,0]
        Disposal_facteur = [0,0,0]
        Skin_facteur = [0,0,0]
        Human_food_facteur = [0,0,0]
        Fat_facteur = [0,0,0]
        Spreading_facteur = [0,0,0]

        for i in range(0,len(facteurs)):
            if facteurs[i][1] == 'Pet food':
                Pet_food_facteur[0] = (float(Pet_food_facteur[0]) + float(facteurs[i][2]))
                Pet_food_facteur[1] = (float(Pet_food_facteur[1]) + float(facteurs[i][3]))
                Pet_food_facteur[2] = (float(Pet_food_facteur[2]) + float(facteurs[i][4]))
            if facteurs[i][1] == 'PAP C3':
                PAP_facteur[0] = (float(PAP_facteur[0]) + float(facteurs[i][2]))
                PAP_facteur[1] = (float(PAP_facteur[1]) + float(facteurs[i][3]))
                PAP_facteur[2] = (float(PAP_facteur[2]) + float(facteurs[i][4]))
            if facteurs[i][1] == 'Gelatin C3':
                Gelatin_facteur[0] = (float(Gelatin_facteur[0]) + float(facteurs[i][2]))
                Gelatin_facteur[1] = (float(Gelatin_facteur[1]) + float(facteurs[i][3]))
                Gelatin_facteur[2] = (float(Gelatin_facteur[2]) + float(facteurs[i][4]))
            if facteurs[i][1] == 'C1-C2 for disposal':
                Disposal_facteur[0] = (float(Disposal_facteur[0]) + float(facteurs[i][2]))
                Disposal_facteur[1] = (float(Disposal_facteur[1]) + float(facteurs[i][3]))
                Disposal_facteur[2] = (float(Disposal_facteur[2]) + float(facteurs[i][4]))
            if facteurs[i][1] == 'Skin tannery C3':
                Skin_facteur[0] = (float(Skin_facteur[0]) + float(facteurs[i][2]))
                Skin_facteur[1] = (float(Skin_facteur[1]) + float(facteurs[i][3]))
                Skin_facteur[2] = (float(Skin_facteur[2]) + float(facteurs[i][4]))
            if facteurs[i][1] == 'Human food':
                Human_food_facteur[0] = (float(Human_food_facteur[0]) + float(facteurs[i][2]))
                Human_food_facteur[1] = (float(Human_food_facteur[1]) + float(facteurs[i][3]))
                Human_food_facteur[2] = (float(Human_food_facteur[2]) + float(facteurs[i][4]))
            if facteurs[i][1] == 'Fat and greaves C3':
                Fat_facteur[0] = (float(Fat_facteur[0]) + float(facteurs[i][2]))
                Fat_facteur[1] = (float(Fat_facteur[1]) + float(facteurs[i][3]))
                Fat_facteur[2] = (float(Fat_facteur[2]) + float(facteurs[i][4]))
            if facteurs[i][1] == 'Spreading/Compost':
                Spreading_facteur[0] = (float(Spreading_facteur[0]) + float(facteurs[i][2]))
                Spreading_facteur[1] = (float(Spreading_facteur[1]) + float(facteurs[i][3]))
                Spreading_facteur[2] = (float(Spreading_facteur[2]) + float(facteurs[i][4]))

        alloc_fact_tri.append(Pet_food_facteur)
        alloc_fact_tri.append(PAP_facteur)
        alloc_fact_tri.append(Gelatin_facteur)
        alloc_fact_tri.append(Disposal_facteur)
        alloc_fact_tri.append(Skin_facteur)
        alloc_fact_tri.append(Human_food_facteur)
        alloc_fact_tri.append(Fat_facteur)
        alloc_fact_tri.append(Spreading_facteur)

        # On remplit le tableau des facteurs d'allocation par destination
        alloc_fact_tri_destinations = ['Pet Food', 'PAP C3', 'Gelatin C3', 'C1-C2 for disposal', 'Skin tannery C3', 'Human food', 'Fat and greaves C3', 'Spreading/Compost']
        alloc_fact_tri_tableau = []
        for i in range(0,len(alloc_fact_tri)):
            alloc_fact_tri_tableau.append([alloc_fact_tri_destinations[i], "%.4f" % alloc_fact_tri[i][0], "%.4f" % alloc_fact_tri[i][1], "%.4f" % alloc_fact_tri[i][2]])
        for k in range(0,len(alloc_fact_tri_tableau)):
            for l in range(0,len(alloc_fact_tri_tableau[0])):
                valeur_numerique_defaut = QtWidgets.QTableWidgetItem(str(alloc_fact_tri_tableau[k][l]))
                self.ui.table_facteurs_par_destination.setItem(k,l,valeur_numerique_defaut)


    # On trace un barchart horizontal comparant les 3 méthodes d'allocation
        facteurs_graphe = ('Allocation biophysique','Allocation massique','Allocation économique') # Liste des coproduits (les bars du barchart)
        # mise en forme des données
        data = []
        for i in range(0,len(alloc_fact_tri)):
            data.append([alloc_fact_tri[i][0],alloc_fact_tri[i][1],alloc_fact_tri[i][2]])

        y_pos = np.arange(len(facteurs_graphe)) # On définit la position du graphe à gauche pour qui soit adaptée aux labels
        plt.rcParams.update({'figure.autolayout': True}) # Le graphe se redimensionnera automatiquement pour permettre l'affichage des labels
        fig = plt.figure(figsize=(ratio_largeur2*16,ratio_hauteur2*8)) # On définit le nom du graphe et sa taille
        ax = fig.add_subplot(111)
        # On définit une liste de couleurs qui seront utilisées pour représenter chaque destination
        colors = ['#C23859', '#C238B5', '#8C38C2', '#386EC2', '#38C2A5', '#98C238', '#C29838', '#C24938']

        # On trace une barre pour chaque facteur d'allocation
        patch_handles = []
        left = np.zeros(len(facteurs_graphe))
        for i, d in enumerate(data):
            patch_handles.append(ax.barh(y_pos, d,  label = liste_destination[i],
                                color=colors[i%len(colors)], align='center', left=left))
            left += d

        # On légende l'axe vertical et on met un titre
        ax.set_yticks(y_pos)
        ax.set_yticklabels(facteurs_graphe)
        ax.set_title("Facteurs d'allocation des coproduits de la viande")

        # On affiche la légende en la positionnant correctement
        plt.legend(bbox_to_anchor=(1.05,1))

        # On affiche la figure sur la fenêtre graphique
        scene = QtWidgets.QGraphicsScene(self)
        self.ui.graphe.scene = scene
        canvas = FigureCanvas(fig)
        #canvas.setGeometry(700, 500, 900, 250)
        canvas.setGeometry(ratio_largeur2*700, ratio_hauteur2*500, ratio_largeur2*900, ratio_hauteur2*250)
        scene.addWidget(canvas)
        self.ui.graphe.setScene(scene)

        self.facteurs = facteurs
        self.alloc_fact_tri = alloc_fact_tri

    def Excel(self):
        # on rappelle la liste des facteurs d'allocations calculés précédemment
        facteurs = self.facteurs
        # on rappelle également la liste de facteurs d'allocation triée par destination
        alloc_fact_tri = self.alloc_fact_tri
        alloc_fact_tri_destinations = ['Pet Food', 'PAP C3', 'Gelatin C3', 'C1-C2 for disposal', 'Skin tannery C3', 'Human food', 'Fat and greaves C3', 'Spreading/Compost']
        # on définit le fichier où sera enregistrée la fiche
        nouvelle_fiche_Excel = QtWidgets.QFileDialog.getSaveFileName(self, "Open File", "", "Excel Files (*.xlsx)");
        path = nouvelle_fiche_Excel[0]
        # on crée un fichier excel xlsx
        wb = Workbook()
        # on sélectionne la première feuille
        ws = wb.active
        # on renomme la feuille
        ws.title = "Données d'entrée"
        # on crée deux autres feuilles
        ws2 = wb.create_sheet("AF by coproduct")
        ws3 = wb.create_sheet("AF by destination")
        # On définit des styles de bordures
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

        # on remplit la feuille 2
        # on nomme les colonnes et on modifie la police des noms de colonne
        ws2.cell(row = 1, column = 3, value = 'Total impact weighting (per coproduct)')
        ws2.cell(row = 1, column = 6, value = 'Allocation Factors (per kg of coproduct)')
        ws2.cell(row = 2, column = 1, value = 'Co-product')
        ws2.cell(row = 2, column = 2, value = 'Destination')
        ws2.cell(row = 2, column = 3, value = 'Biophysical allocation')
        ws2.cell(row = 2, column = 4, value = 'Mass allocation')
        ws2.cell(row = 2, column = 5, value = 'Economic allocation')
        ws2.cell(row = 2, column = 6, value = 'Biophysical allocation')
        ws2.cell(row = 2, column = 7, value = 'Mass allocation')
        ws2.cell(row = 2, column = 8, value = 'Economic allocation')
        for i in range(1,9):
            ws2.cell(row = 1, column = i).font = Font(bold=True)
            ws2.cell(row = 1, column = i).font = Font(size=14)
            ws2.cell(row = 2, column = i).font = Font(bold=True)
            ws2.cell(row = 2, column = i).font = Font(size=14)
        ws2.merge_cells('C1:E1')
        ws2.merge_cells('F1:H1')
        # on remplit les colonnes et on trace les bordures
        for i in range(2,len(facteurs)+2):
            for j in range(1,len(facteurs[0])+1):
                ws2.cell(row = i+1, column = j, value = facteurs[i-2][j-1])
                ws2.cell(row = i+1, column = j).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws2.cell(row = 1, column = j).border = Border(top=thin, left=thin, right=thin, bottom=double)
                ws2.cell(row = 2, column = j).border = Border(top=thin, left=thin, right=thin, bottom=double)
        # on ajuste la taille des colonnes
        l0 = []
        l1 = []
        for i in range(0,len(facteurs)):
            l0.append(len(facteurs[i][0]))
            l1.append(len(facteurs[i][1]))
            m0 = max(l0)
            m1 = max(l1)
        ws2.column_dimensions['A'].width = m0
        ws2.column_dimensions['B'].width = m1
        ws2.column_dimensions['C'].width = 24
        ws2.column_dimensions['D'].width = 18
        ws2.column_dimensions['E'].width = 23
        ws2.column_dimensions['F'].width = 24
        ws2.column_dimensions['G'].width = 18
        ws2.column_dimensions['H'].width = 23

        # on remplit la feuille 3
        # on nomme les colonnes et on modifie la police des noms de colonne
        ws3.cell(row = 1, column = 1, value = 'Destination')
        ws3.cell(row = 1, column = 2, value = 'Biophysical allocation')
        ws3.cell(row = 1, column = 3, value = 'Mass allocation')
        ws3.cell(row = 1, column = 4, value = 'Economic allocation')
        for i in range(1,5):
            ws3.cell(row = 1, column = i).font = Font(bold=True)
            ws3.cell(row = 1, column = i).font = Font(size=14)
        # on remplit les colonnes
        for i in range(1,len(alloc_fact_tri)+1):
            for j in range(2,len(alloc_fact_tri[0])+2):
                ws3.cell(row = i+1, column = j, value = alloc_fact_tri[i-1][j-2])
                ws3.cell(row = i+1, column = j).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws3.cell(row = 1, column = j).border = Border(top=thin, left=thin, right=thin, bottom=double)
        for i in range(1,len(alloc_fact_tri_destinations)+1):
            ws3.cell(row = i+1, column = 1, value = alloc_fact_tri_destinations[i-1])
            ws3.cell(row = i+1, column = 1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            ws3.cell(row = 1, column = 1).border = Border(top=thin, left=thin, right=thin, bottom=double)
        # on ajuste la taille des colonnes
        l2 = []
        for i in range(0,len(alloc_fact_tri_destinations)):
            l2.append(len(alloc_fact_tri_destinations[i]))
            m2 = max(l2)
        ws3.column_dimensions['A'].width = m2
        ws3.column_dimensions['B'].width = 24
        ws3.column_dimensions['C'].width = 18
        ws3.column_dimensions['D'].width = 23
        # on trace un barchart horizontal
        values = Reference(ws3, min_col=1, min_row=2, max_col=4, max_row=len(alloc_fact_tri_destinations)+1)
        cats = Reference(ws3, min_row=1, min_col=2, max_col=9)
        chart = BarChart()
        chart.title = "Allocation Factors by Destination"
        chart.add_data(values, titles_from_data=True, from_rows=True)
        chart.set_categories(cats)
        chart.type = "bar"
        chart.style = 2
        chart.grouping = "percentStacked"
        chart.overlap = 100
        ws3.add_chart(chart, "B12")

        # On récupère la liste qui représente l'individu sélectionné et on va l'afficher dans l'onglet "Données d'entrée" du fichier Excel
        data_individu = self.data_individu
        labels_colonne = ['Espèce','Race','Catégorie',"Mode d'élevage",'Co-produits','Destination','Group of tissues','Pourcentage du poids total',
                         'Eau (%)','Matière sèche  (%)','Lipides  (%)','Protéines (%)','Gomperts coefficient','EBW0','EBWm',
                         'EBWsl','FAT0','FATm','FATmax','Rwp','Z1prot','Z2lip','Act_Coeff','Paramètre énergétique','Valeur économique','Rendement Carcasse']
        for i in range(0,len(data_individu)):
            for j in range(0,len(data_individu[0])):
                ws.cell(row = i+2, column = j+1 , value = data_individu[i][j])
                ws.cell(row = 1, column = j+1, value = labels_colonne[j])
                ws.cell(row = i+2, column = j+1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(row = 1, column = j+1).border = Border(top=thin, left=thin, right=thin, bottom=double)
        # On modifie la police des titres de colonne
        for i in range(1,30):
            ws.cell(row = 1, column = i).font = Font(bold=True)
            ws.cell(row = 1, column = i).font = Font(size=14)
        # on ajuste la taille des colonnes
        l1 = []
        l4 = []
        l5 = []
        for i in range(0,len(data_individu)):
            l1.append(len(data_individu[i][1]))
            l1.append(6)
            l4.append(len(data_individu[i][4]))
            l5.append(len(data_individu[i][5]))
            m1 = max(l1)
            m4 = max(l4)
            m5 = max(l5)
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = m1
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = m4
        ws.column_dimensions['F'].width = m5
        ws.column_dimensions['G'].width = 19
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 21
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 17
        ws.column_dimensions['M'].width = 24
        ws.column_dimensions['N'].width = 10
        ws.column_dimensions['O'].width = 10
        ws.column_dimensions['P'].width = 10
        ws.column_dimensions['Q'].width = 10
        ws.column_dimensions['R'].width = 10
        ws.column_dimensions['S'].width = 10
        ws.column_dimensions['T'].width = 10
        ws.column_dimensions['U'].width = 10
        ws.column_dimensions['V'].width = 10
        ws.column_dimensions['W'].width = 15
        ws.column_dimensions['X'].width = 27
        ws.column_dimensions['Y'].width = 24
        ws.column_dimensions['Z'].width = 24

        # on enregistre le fichier
        wb.save(path)


    # On remplit la combobox2 en fonction de la valeur sélectionnée dans la combobox1
    def Combobox4(self):
        self.ui.combobox5.clear()
        liste_especes = self.liste_especes
        liste_finale = self.liste_finale
        liste_races = []

        for i in range(0,len(liste_especes)):
            if self.ui.combobox4.currentText() == liste_especes[i]:
                for k in range(0, len(liste_finale)):
                    for j in range(0, len(liste_finale[k])):
                        if liste_finale[k][j][0] == liste_especes[i]:
                            liste_races.append(liste_finale[k][j][1])
        self.liste_races = list(set(liste_races))

        for i in range(0,len(self.liste_races)):
            self.ui.combobox5.addItem(self.liste_races[i])

    # On remplit la combobox3 en fonction de la valeur sélectionnée dans la combobox2
    def Combobox5(self):
        self.ui.combobox6.clear()
        liste_especes = self.liste_especes
        liste_finale = self.liste_finale
        liste_races = self.liste_races
        liste_mois = []

        for o in range(0,len(liste_especes)):
            if self.ui.combobox4.currentText() == liste_especes[o]:
                for i in range(0,len(liste_races)):
                    if self.ui.combobox5.currentText() == liste_races[i]:
                        for k in range(0, len(liste_finale)):
                            for j in range(0, len(liste_finale[k])):
                                if liste_finale[k][j][0] == liste_especes[o] and liste_finale[k][j][1] == liste_races[i]:
                                    liste_mois.append(liste_finale[k][j][2])
        self.liste_mois = list(set(liste_mois))

        for i in range(0,len(self.liste_mois)):
            self.ui.combobox6.addItem(self.liste_mois[i])


    # On remplit la combobox7 en fonction de la valeur sélectionnée dans la combobox3
    def Combobox6(self):
        self.ui.combobox8.clear()
        liste_especes = self.liste_especes
        liste_finale = self.liste_finale
        liste_races = self.liste_races
        liste_mois = self.liste_mois
        liste_elevage = []

        for o in range(0,len(liste_especes)):
            if self.ui.combobox4.currentText() == liste_especes[o]:
                for p in range(0,len(liste_races)):
                    if self.ui.combobox5.currentText() == liste_races[p]:
                        for i in range(0, len(liste_mois)):
                            if self.ui.combobox6.currentText() == liste_mois[i]:
                                for k in range(0, len(liste_finale)):
                                    for j in range(0, len(liste_finale[k])):
                                        if liste_finale[k][j][0] == liste_especes[o] and liste_finale[k][j][1] == liste_races[p] and liste_finale[k][j][2] == liste_mois[i]:
                                            liste_elevage.append(liste_finale[k][j][3])
        liste_elevage = list(set(liste_elevage))

        for i in range(0,len(liste_elevage)):
            self.ui.combobox8.addItem(liste_elevage[i])

    # Creation de la fonction ajouter qui permet d'ajouter un individu au *.csv
    def Ajouter(self):
        liste_finale = self.liste_finale
        Error_Item = False
        for i in range(0,len(liste_finale)):
            if (self.ui.texte_espece.toPlainText() == liste_finale[i][0][0] and
                self.ui.texte_race.toPlainText() == liste_finale[i][0][1] and
                self.ui.texte_mois.toPlainText() == liste_finale[i][0][2] and
                self.ui.texte_elevage.toPlainText() == liste_finale[i][0][3]):
                Error_Item = True
                break

        # Test et renvoi d'un message s'il manque un renseignement sur l'individu (espèce, race ou mois d'abattage)
        if (self.ui.texte_espece.toPlainText() == '' or
            (self.ui.texte_espece.toPlainText() != 'Bovine' and
            self.ui.texte_espece.toPlainText() != 'Ovine' and
            self.ui.texte_espece.toPlainText() != 'Calf')):
            Error_Espece = True
        else:
            Error_Espece = False

        if (self.ui.texte_race.toPlainText() == ''):
            Error_Race = True
        else:
            Error_Race = False

        if (self.ui.texte_mois.toPlainText() == ''):
            Error_Category = True
        else:
            Error_Category = False

        if (self.ui.texte_elevage.toPlainText() == ''):
            Error_Rearing = True
        else:
            Error_Rearing = False

        # Test et renvoi d'un message si la liste des coproduits n'est pas valide
        if (self.ui.table4.item(0,0) == None):
            Error_Coproduct = True
        else:
            Error_Coproduct = False

        # Test et renvoi d'un message si des données de paramètre manquantes
        if ((self.ui.table5.item(0,0) == None) or
            (not isfloat(self.ui.table5.item(0,0).text()))):
            Error_Gompertz = True
        else:
            Error_Gompertz = False

        if ((self.ui.table5.item(1,0) == None) or
            (not isfloat(self.ui.table5.item(1,0).text()))):
            Error_EBW0 = True
        else:
            Error_EBW0 = False

        if ((self.ui.table5.item(2,0) == None) or
            (not isfloat(self.ui.table5.item(2,0).text()))):
            Error_EBWm = True
        else:
            Error_EBWm = False

        if ((self.ui.table5.item(3,0) == None) or
            (not isfloat(self.ui.table5.item(3,0).text()))):
            Error_EBWsl = True
        else:
            Error_EBWsl = False

        if ((self.ui.table5.item(4,0) == None) or
            (not isfloat(self.ui.table5.item(4,0).text()))):
            Error_FAT0 = True
        else:
            Error_FAT0 = False

        if ((self.ui.table5.item(5,0) == None) or
            (not isfloat(self.ui.table5.item(5,0).text()))):
            Error_FATm = True
        else:
            Error_FATm = False

        if ((self.ui.table5.item(6,0) == None) or
            (not isfloat(self.ui.table5.item(6,0).text()))):
            Error_FATmax = True
        else:
            Error_FATmax = False

        if ((self.ui.table5.item(7,0) == None) or
            (not isfloat(self.ui.table5.item(7,0).text()))):
            Error_Rwp = True
        else:
            Error_Rwp = False

        if ((self.ui.table5.item(8,0) == None) or
            (not isfloat(self.ui.table5.item(8,0).text()))):
            Error_z1 = True
        else:
            Error_z1 = False

        if ((self.ui.table5.item(9,0) == None) or
            (not isfloat(self.ui.table5.item(9,0).text()))):
            Error_z2 = True
        else:
            Error_z2 = False

        if ((self.ui.table5.item(10,0) == None) or
            (not isfloat(self.ui.table5.item(10,0).text()))):
            Error_act_coeff = True
        else:
            Error_act_coeff = False

        if ((self.ui.table5.item(11,0) == None) or
            (not isfloat(self.ui.table5.item(11,0).text()))):
            Error_Carcass_Yield = True
        else:
            Error_Carcass_Yield = False

        Error_list = [Error_Item, Error_Espece, Error_Race, Error_Category, Error_Rearing, Error_Coproduct, Error_Gompertz, Error_EBW0, Error_EBWm, Error_EBWsl, Error_FAT0, Error_FATm, Error_FATmax, Error_Rwp, Error_z1, Error_z2, Error_act_coeff, Error_Carcass_Yield]
        Error_name = ['Error_Item','Error_Espece', 'Error_Race', 'Error_Category', 'Error_Rearing', 'Error_Coproduct', 'Error_Gompertz', 'Error_EBW0', 'Error_EBWm', 'Error_EBWsl', 'Error_FAT0', 'Error_FATm', 'Error_FATmax', 'Error_Rwp', 'Error_z1', 'Error_z2', 'Error_act_coeff', 'Error_Carcass_Yield']
        for i in range(0,len(Error_list)):
            if Error_list[i] == True:
                msg_espece = QtWidgets.QMessageBox()
                msg_espece.setWindowTitle(Title.get(Error_name[i]))
                msg_espece.setIcon(msg_espece.Warning)
                msg_espece.setText(Text.get(Error_name[i]))
                msg_espece.setDetailedText(Details.get(Error_name[i]))
                Bouton_Retour_espece = msg_espece.addButton("Cancel", msg_espece.YesRole)
                msg_espece.exec_()
                return

        # Test et renvoi d'un message si un co-produit saisi est incomplet - possibilité d'utiliser des valeurs par défaut
        for i in range(0,70):
            if ((self.ui.table4.item(i,0) != None) and
               ((self.ui.table4.cellWidget(i,1).currentText() == 'To be specified') or
               (self.ui.table4.cellWidget(i,2).currentText() == 'To be specified') or
               (self.ui.table4.item(i,3) == None) or
               (self.ui.table4.item(i,4) == None) or
               (self.ui.table4.item(i,5) == None) or
               (self.ui.table4.item(i,6) == None) or
               (self.ui.table4.item(i,7) == None))):

                msg_missing_data = QtWidgets.QMessageBox()
                msg_missing_data.setWindowTitle("Missing data")
                msg_missing_data.setIcon(msg_missing_data.Warning)
                msg_missing_data.setText("Some data are missing")
                msg_missing_data.setInformativeText("You can use default value (see details)")
                msg_missing_data.setDetailedText("Default values will be: \n\n Destination : C1-C2 for disposal \n Group of Tissus : Others \n Toutes les valeurs numériques : 0")
                Bouton_Defaut_missing_data = msg_missing_data.addButton("Use default values", msg_missing_data.YesRole)
                Bouton_Saisie_missing_data = msg_missing_data.addButton("Manual input", msg_missing_data.NoRole)
                msg_missing_data.exec_()
                if msg_missing_data.clickedButton() == Bouton_Defaut_missing_data:
                    for i in range(0,70):
                        if ((self.ui.table4.item(i,0) != None) and
                           (self.ui.table4.cellWidget(i,1).currentText() == 'To be specified')):
                            self.ui.table4.cellWidget(i,1).setCurrentText('C1-C2 for disposal')
                        if ((self.ui.table4.item(i,0) != None) and
                           (self.ui.table4.cellWidget(i,2).currentText() == 'To be specified')):
                            self.ui.table4.cellWidget(i,2).setCurrentText('Others')
                        if (self.ui.table4.item(i,0) != None) and (self.ui.table4.item(i,3) == None):
                            valeur_numerique_defaut = QtWidgets.QTableWidgetItem('0')
                            self.ui.table4.setItem(i,3,valeur_numerique_defaut)
                        if (self.ui.table4.item(i,0) != None) and (self.ui.table4.item(i,4) == None):
                            valeur_numerique_defaut = QtWidgets.QTableWidgetItem('0')
                            self.ui.table4.setItem(i,4,valeur_numerique_defaut)
                        if (self.ui.table4.item(i,0) != None) and (self.ui.table4.item(i,5) == None):
                            valeur_numerique_defaut = QtWidgets.QTableWidgetItem('0')
                            self.ui.table4.setItem(i,5,valeur_numerique_defaut)
                        if (self.ui.table4.item(i,0) != None) and (self.ui.table4.item(i,6) == None):
                            valeur_numerique_defaut = QtWidgets.QTableWidgetItem('0')
                            self.ui.table4.setItem(i,6,valeur_numerique_defaut)
                        if (self.ui.table4.item(i,0) != None) and (self.ui.table4.item(i,7) == None):
                            valeur_numerique_defaut = QtWidgets.QTableWidgetItem('0')
                            self.ui.table4.setItem(i,7,valeur_numerique_defaut)
                    return

                if msg_missing_data.clickedButton() == Bouton_Saisie_missing_data:
                    return



        # Test et renvoi d'un message s'il manque des données de paramètre énergétique - possibilité d'utiliser des valeurs par défaut
        for i in range(0,4):
            if ((self.ui.table6.item(i,0) == None) or
                (not isfloat(self.ui.table6.item(i,0).text()))):
                msg_no_ener_par = QtWidgets.QMessageBox()
                msg_no_ener_par.setWindowTitle("Energetic data missing")
                msg_no_ener_par.setIcon(msg_no_ener_par.Warning)
                msg_no_ener_par.setText("Some energetic parameters are missing")
                msg_no_ener_par.setInformativeText("You can use default value (see details)")
                msg_no_ener_par.setDetailedText("Default values will be: \n\n Whole Body : 0.613 \n Carcass : 0.354 \n GIT : 4.024 \n Liver : 2.013 \n Others : 0.343")
                Bouton_Defaut_no_ener_par = msg_no_ener_par.addButton("Use default values", msg_no_ener_par.YesRole)
                Bouton_Saisie_no_ener_par = msg_no_ener_par.addButton("Manual input", msg_no_ener_par.NoRole)
                msg_no_ener_par.exec_()
                if msg_no_ener_par.clickedButton() == Bouton_Defaut_no_ener_par:
                    for i in range(0,4):
                        if ((self.ui.table6.item(0,0) == None) or
                            (not isfloat(self.ui.table6.item(0,0).text()))):
                            valeur_energie_0 = QtWidgets.QTableWidgetItem('0.613')
                            self.ui.table6.setItem(0,0,valeur_energie_0)
                        if ((self.ui.table6.item(1,0) == None) or
                            (not isfloat(self.ui.table6.item(1,0).text()))):
                            valeur_energie_1 = QtWidgets.QTableWidgetItem('0.354')
                            self.ui.table6.setItem(1,0,valeur_energie_1)
                        if ((self.ui.table6.item(2,0) == None) or
                            (not isfloat(self.ui.table6.item(2,0).text()))):
                            valeur_energie_2 = QtWidgets.QTableWidgetItem('4.024')
                            self.ui.table6.setItem(2,0,valeur_energie_2)
                        if ((self.ui.table6.item(3,0) == None) or
                            (not isfloat(self.ui.table6.item(3,0).text()))):
                            valeur_energie_3 = QtWidgets.QTableWidgetItem('2.013')
                            self.ui.table6.setItem(3,0,valeur_energie_3)
                        if ((self.ui.table6.item(4,0) == None) or
                            (not isfloat(self.ui.table6.item(4,0).text()))):
                            valeur_energie_4 = QtWidgets.QTableWidgetItem('0.343')
                            self.ui.table6.setItem(4,0,valeur_energie_4)
                    return

                if msg_no_ener_par.clickedButton() == Bouton_Saisie_no_ener_par:
                    return

            # Test et renvoi d'un message si la somme des poids des co-produits ne fait pas 100%
            new_BW = []
            for i in range(0,70):
                if self.ui.table4.item(i,0) != None:
                    new_BW.append(float(self.ui.table4.item(i,3).text()))
            totalBW = sum(new_BW)
            if (totalBW - 100) > 0.01 or (totalBW - 100) < -0.01:
                msg_totalBW = QtWidgets.QMessageBox()
                msg_totalBW.setWindowTitle("Total mass is not 100%")
                msg_totalBW.setIcon(msg_totalBW.Warning)
                msg_totalBW.setText("The sum of the mass of the coproducts is not equal to 100% but to : " + str("%.2f" % totalBW) + "%")
                msg_totalBW.setInformativeText("Do you want to use automatic adjustment ? (see details)")
                msg_totalBW.setDetailedText("The mass fraction of each coproduct will be multiplicated by : \n\n the ratio between 100% and current total mass \n\n Automatic adjustment is not recommended if the current total is far from 100%. Better check first if no coproduct have been forgotten")
                Bouton_Defaut_BW = msg_totalBW.addButton("Automatic adjustment", msg_totalBW.YesRole)
                Bouton_Saisie_BW = msg_totalBW.addButton("Manual adjustment", msg_totalBW.NoRole)
                msg_totalBW.exec_()
                if msg_totalBW.clickedButton() == Bouton_Defaut_BW:
                    for i in range(0,70):
                        if self.ui.table4.item(i,0) != None:
                            data_BW0 = QtWidgets.QTableWidgetItem("%.4f" % (float(new_BW[i])*(100/totalBW)))
                            self.ui.table4.setItem(i,3,data_BW0)
                    return
                if msg_totalBW.clickedButton() == Bouton_Saisie_BW:
                    return

            # teste si les rendements carcasse théorique et calculé sont égaux et sinon, propose un ajustement
            BW_carcass = []
            BW_non_carcass = []
            for i in range(0,70):
                if self.ui.table4.item(i,0) != None:
                    if self.ui.table4.cellWidget(i,2).currentText() == 'Carcass':
                        BW_carcass.append(float(self.ui.table4.item(i,3).text()))
                    else:
                        BW_non_carcass.append(float(self.ui.table4.item(i,3).text()))
            carcass_yield_calcul = sum(BW_carcass)/100
            carcass_yield_corr = carcass_yield_calcul/float(self.ui.table5.item(11,0).text())
            if (carcass_yield_calcul - float(self.ui.table5.item(11,0).text())) > 0.01 or (carcass_yield_calcul - float(self.ui.table5.item(11,0).text())) < -0.01:
                msg_carcass_yield_diff = QtWidgets.QMessageBox()
                msg_carcass_yield_diff.setWindowTitle("Calculated carcass yield is different from entered carcass yield")
                msg_carcass_yield_diff.setIcon(msg_carcass_yield_diff.Warning)
                msg_carcass_yield_diff.setText("Calculated carcass yield is different from entered carcass yield")
                msg_carcass_yield_diff.setInformativeText("Do you want to use automatic adjustment ? (see details)")
                msg_carcass_yield_diff.setDetailedText("The mass of the coproducts will be adjusted to the carcass yield entered in table")
                Bouton_Defaut_carcass_yield_diff = msg_carcass_yield_diff.addButton("Automatic adjustment", msg_carcass_yield_diff.YesRole)
                Bouton_Saisie_carcass_yield_diff = msg_carcass_yield_diff.addButton("Manual adjustment", msg_carcass_yield_diff.NoRole)
                msg_carcass_yield_diff.exec_()
                if msg_carcass_yield_diff.clickedButton() == Bouton_Defaut_carcass_yield_diff:
                    for i in range(0,70):
                        if self.ui.table4.item(i,0) != None:
                            if self.ui.table4.cellWidget(i,2).currentText() == 'Carcass':
                                valeur_BW = QtWidgets.QTableWidgetItem(str(float(self.ui.table4.item(i,3).text())/carcass_yield_corr))
                                self.ui.table4.setItem(i,3,valeur_BW)
                            else:
                                valeur_BW = QtWidgets.QTableWidgetItem(str(float(self.ui.table4.item(i,3).text())*carcass_yield_corr))
                                self.ui.table4.setItem(i,3,valeur_BW)
                    return
                if msg_carcass_yield_diff.clickedButton() == Bouton_Saisie_carcass_yield_diff:
                    return


        else:
            #Creation de listes qui vont représenter le nouvel individu
            new_coproduct = []
            new_destination = []
            new_group_of_tissues = []
            new_BW = []
            new_water = []
            new_dry_matter = []
            new_lipid = []
            new_protein = []
            new_individu = []
            new_valeur_economique = []


            # Remplissage des listes avec les valeurs entrées par l'utilisateur dans le table4
            for i in range(0,70):
                if self.ui.table4.item(i,0) != None:
                    new_coproduct.append(self.ui.table4.item(i,0).text())
                    new_destination.append(self.ui.table4.cellWidget(i,1).currentText())
                    new_group_of_tissues.append(self.ui.table4.cellWidget(i,2).currentText())
                    new_BW.append(float(self.ui.table4.item(i,3).text()))
                    new_water.append(float(self.ui.table4.item(i,4).text()))
                    new_dry_matter.append(float(self.ui.table4.item(i,5).text()))
                    new_lipid.append(float(self.ui.table4.item(i,6).text()))
                    new_protein.append(float(self.ui.table4.item(i,7).text()))
                    new_valeur_economique.append(float(self.ui.table4.item(i,8).text()))

            # Remplissage des autres éléments du *.csv avec les valeurs de l'utilisateur
            new_espece = self.ui.texte_espece.toPlainText()
            new_race = self.ui.texte_race.toPlainText()
            new_mois = self.ui.texte_mois.toPlainText()
            new_elevage = self.ui.texte_elevage.toPlainText()
            new_Gompertz = float(self.ui.table5.item(0,0).text())
            new_EBW0 = float(self.ui.table5.item(1,0).text())
            new_EBWm = float(self.ui.table5.item(2,0).text())
            new_EBWsl = float(self.ui.table5.item(3,0).text())
            new_FAT0 = float(self.ui.table5.item(4,0).text())
            new_FATm = float(self.ui.table5.item(5,0).text())
            new_FATmax = float(self.ui.table5.item(6,0).text())
            new_Rwp = float(self.ui.table5.item(7,0).text())
            new_z1 = float(self.ui.table5.item(8,0).text())
            new_z2 = float(self.ui.table5.item(9,0).text())
            new_activity = float(self.ui.table5.item(10,0).text())
            new_rendement_carcasse = float(self.ui.table5.item(11,0).text())

            new_parametre_energie = []
            for i in range(0,len(new_coproduct)):
                if new_group_of_tissues[i] == 'whole body':
                    new_parametre_energie.append(float(self.ui.table6.item(0,0).text()))
                if new_group_of_tissues[i] == 'Carcass':
                    new_parametre_energie.append(float(self.ui.table6.item(1,0).text()))
                if new_group_of_tissues[i] == 'GIT':
                    new_parametre_energie.append(float(self.ui.table6.item(2,0).text()))
                if new_group_of_tissues[i] == 'Liver':
                    new_parametre_energie.append(float(self.ui.table6.item(3,0).text()))
                if new_group_of_tissues[i] == 'Others':
                    new_parametre_energie.append(float(self.ui.table6.item(4,0).text()))


            new_individu = []
            # Remplissage de la liste représentant le nouvel individu grâce aux listes créées précédemment
            for i in range(0,len(new_coproduct)):
                (new_individu.append([new_espece,new_race,new_mois,new_elevage,new_coproduct[i],
                 new_destination[i], new_group_of_tissues[i], new_BW[i],new_water[i],
                 new_dry_matter[i],new_lipid[i], new_protein[i], new_Gompertz,
                 new_EBW0,new_EBWm,new_EBWsl,new_FAT0,new_FATm,new_FATmax,
                 new_Rwp,new_z1,new_z2,new_activity,new_parametre_energie[i],
                 new_valeur_economique[i],new_rendement_carcasse]))


        # On ajoute l'individu à la liste_finale
        liste_finale = self.liste_finale[:]
        liste_finale.append(new_individu)
        titres = ['Species','Breed','Category',"Rearing Method",'Co-product','Destination','Group of tissues','BW (%)',
                 'Water (%)','DM  (%)','Lipids  (%)','Proteins (%)','Gompertz Coeff.','EBW0','EBWm',
                 'EBWsl','FAT0','FATm','FATsl','Rwp','Z1prot','Z2lip','Act. Coeff.','Energy for Maint.','Value','Carcass Yield']
        #liste_finale.insert(0,titres)

        # On choisit le fichier dans lequel on va enregistrer le nouvel individu (soit un fichier existant, soit un nouveau fichier)
        nouvelle_base = QtWidgets.QFileDialog.getSaveFileName(self, "Open File", "", "CSV Files (*.csv)");
        nouvelle_base = nouvelle_base[0]

        # On ouvre le fichier en mode écriture
        bdd_ouverte = open(nouvelle_base, 'w', newline="", encoding = 'utf-8')

        # On écrit la liste finale dans le fichier ouvert
        writer_bdd = csv.writer(bdd_ouverte,delimiter = ";")

        writer_bdd.writerow(titres)
        for j in range(0,len(liste_finale)):
            for i in range(0,len(liste_finale[j])):
                writer_bdd.writerow(liste_finale[j][i])


    def Modele(self):
        for i in range(0,70):
            self.ui.table4.setItem(i,0,None)
            data_destination = 'To be specified'
            self.ui.table4.cellWidget(i,1).setCurrentText(data_destination)
            data_group_of_tissues = 'To be specified'
            self.ui.table4.cellWidget(i,2).setCurrentText(data_group_of_tissues)
            self.ui.table4.setItem(i,3,None)
            self.ui.table4.setItem(i,4,None)
            self.ui.table4.setItem(i,5,None)
            self.ui.table4.setItem(i,6,None)
            self.ui.table4.setItem(i,7,None)
            self.ui.table4.setItem(i,8,None)

        self.ui.texte_espece.setText(self.ui.combobox4.currentText())
        self.ui.texte_race.setText(self.ui.combobox5.currentText())
        self.ui.texte_mois.setText(self.ui.combobox6.currentText())
        self.ui.texte_elevage.setText(self.ui.combobox8.currentText())

        # On crée une liste représentant l'individu dès qu'on change la valeur de la combobox3
        individu = []
        individu.append(self.ui.combobox4.currentText())
        individu.append(self.ui.combobox5.currentText())
        individu.append(self.ui.combobox6.currentText())
        individu.append(self.ui.combobox8.currentText())

        # On crée une liste qui représente l'individu et ses données associées et on l'affiche dans le tableau
        data_individu = []

        if self.ui.combobox8.count() != 0 :
            liste_finale = self.liste_finale
            for i in range(0,len(liste_finale)):
                for j in range(0, len(liste_finale[i])):
                    if (liste_finale[i][j][0] == individu[0] and
                        liste_finale[i][j][1] == individu[1] and
                        liste_finale[i][j][2] == individu[2] and
                        liste_finale[i][j][3] == individu[3]) :
                        data_individu.append(liste_finale[i][j])

            for i in range(0,len(data_individu)):
                data_coproduit = QtWidgets.QTableWidgetItem(data_individu[i][4])
                self.ui.table4.setItem(i,0,data_coproduit)
                data_destination = data_individu[i][5]
                self.ui.table4.cellWidget(i,1).setCurrentText(data_destination)
                data_group_of_tissues = data_individu[i][6]
                self.ui.table4.cellWidget(i,2).setCurrentText(data_group_of_tissues)
                data_BW0 = QtWidgets.QTableWidgetItem("%.4f" % float(data_individu[i][7]))
                self.ui.table4.setItem(i,3,data_BW0)
                data_water = QtWidgets.QTableWidgetItem(data_individu[i][8])
                self.ui.table4.setItem(i,4,data_water)
                data_dry_matter = QtWidgets.QTableWidgetItem(data_individu[i][9])
                self.ui.table4.setItem(i,5,data_dry_matter)
                data_lipides = QtWidgets.QTableWidgetItem(data_individu[i][10])
                self.ui.table4.setItem(i,6,data_lipides)
                data_proteines = QtWidgets.QTableWidgetItem(data_individu[i][11])
                self.ui.table4.setItem(i,7,data_proteines)
                data_Gompertz = QtWidgets.QTableWidgetItem(data_individu[0][12])
                self.ui.table5.setItem(0,0,data_Gompertz)
                data_economique = QtWidgets.QTableWidgetItem(data_individu[i][24])
                self.ui.table4.setItem(i,8,data_economique)
                data_EBW0 = QtWidgets.QTableWidgetItem(data_individu[0][13])
                self.ui.table5.setItem(1,0,data_EBW0)
                data_EBWm = QtWidgets.QTableWidgetItem(data_individu[0][14])
                self.ui.table5.setItem(2,0,data_EBWm)
                data_EBWsl = QtWidgets.QTableWidgetItem(data_individu[0][15])
                self.ui.table5.setItem(3,0,data_EBWsl)
                data_FAT0 = QtWidgets.QTableWidgetItem(data_individu[0][16])
                self.ui.table5.setItem(4,0,data_FAT0)
                data_FATm = QtWidgets.QTableWidgetItem(data_individu[0][17])
                self.ui.table5.setItem(5,0,data_FATm)
                data_FATmax = QtWidgets.QTableWidgetItem(data_individu[0][18])
                self.ui.table5.setItem(6,0,data_FATmax)
                data_Rwp = QtWidgets.QTableWidgetItem(data_individu[0][19])
                self.ui.table5.setItem(7,0,data_Rwp)
                data_z1 = QtWidgets.QTableWidgetItem(data_individu[0][20])
                self.ui.table5.setItem(8,0,data_z1)
                data_z2 = QtWidgets.QTableWidgetItem(data_individu[0][21])
                self.ui.table5.setItem(9,0,data_z2)
                data_activity = QtWidgets.QTableWidgetItem(data_individu[0][22])
                self.ui.table5.setItem(10,0,data_activity)
                data_rendement_carcasse = QtWidgets.QTableWidgetItem(data_individu[0][25])
                self.ui.table5.setItem(11,0,data_rendement_carcasse)
                data_energy_maintenance = QtWidgets.QTableWidgetItem(data_individu[i][23])
                for j in range(0,5):
                    if self.ui.table6.verticalHeaderItem(j).text() == data_individu[i][6] :
                        self.ui.table6.setItem(j,0, data_energy_maintenance)

        else:
            msg_no_database_loaded = QtWidgets.QMessageBox()
            msg_no_database_loaded.setWindowTitle("No database loaded")
            msg_no_database_loaded.setIcon(msg_no_database_loaded.Warning)
            msg_no_database_loaded.setText("No database loaded")
            msg_no_database_loaded.setInformativeText("To load a database, go back to - Open Database -")
            Bouton_Retour_no_database = msg_no_database_loaded.addButton("Cancel", msg_no_database_loaded.YesRole)
            msg_no_database_loaded.exec_()
            if msg_no_database_loaded.clickedButton() == Bouton_Retour_no_database:
                return

# Un clic sur le bouton Effacer permet de réinitialiser toutes les valeurs de la page
    def Effacer(self):
        for i in range(0,70):
            self.ui.table4.setItem(i,0,None)
            data_destination = 'To be specified'
            self.ui.table4.cellWidget(i,1).setCurrentText(data_destination)
            data_group_of_tissues = 'To be specified'
            self.ui.table4.cellWidget(i,2).setCurrentText(data_group_of_tissues)
            self.ui.table4.setItem(i,3,None)
            self.ui.table4.setItem(i,4,None)
            self.ui.table4.setItem(i,5,None)
            self.ui.table4.setItem(i,6,None)
            self.ui.table4.setItem(i,7,None)
            self.ui.table4.setItem(i,8,None)
        for i in range(0,12):
            self.ui.table5.setItem(i,0,None)
        for i in range(0,5):
            self.ui.table6.setItem(i,0,None)
        self.ui.texte_espece.setText(None)
        self.ui.texte_race.setText(None)
        self.ui.texte_mois.setText(None)
        self.ui.texte_elevage.setText(None)






if __name__ == "__main__":
    appctxt = ApplicationContext()       # 1. Instantiate ApplicationContext
    app = QtWidgets.QApplication(sys.argv)
    w = AppWindow()
    w.show()
    exit_code = appctxt.app.exec_()      # 2. Invoke appctxt.app.exec_()
    sys.exit(exit_code)
